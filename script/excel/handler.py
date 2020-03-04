# -*- coding: utf-8 -*-

from os.path import exists, splitext
from openpyxl import Workbook, load_workbook as XlsxWorkbook

from script.excel.adapter import BookAdapter, XlsSheetAdapter, XlsxSheetAdapter
from script.excel.support import XlsWorkbook
from script.excel.wrapper import SheetWrapper, EmptyWrapper


def iworkbook(path=None) -> BookAdapter:
    """Load workbook from hard disk or create new workbook."""
    # Create new workbook in xlsx format.
    if not path:
        workbook = Workbook()
        decorate = EmptyWrapper(XlsxSheetAdapter(workbook.active))
        return BookAdapter(workbook, [decorate])

    if not exists(path): raise FileExistsError(
        'Workbook file {} does not exist.'.format(path)
    )

    extension = splitext(path)[1]

    # Adapt xlsx workbook and worksheets.
    if extension == '.xlsx':
        workbook = XlsxWorkbook(path)
        allsheet = [
            SheetWrapper(adapter) for adapter in [
                XlsxSheetAdapter(worksheet) for worksheet in workbook
            ]
            if not adapter.empty()
        ]

    # Adapt xls workbook and worksheets.
    elif extension == '.xls':
        workbook = XlsWorkbook(path)
        allsheet = [
            SheetWrapper(adapter) for adapter in [
                XlsSheetAdapter(workbook, index)
                for index in range(workbook.readable.nsheets)
            ]
            if not adapter.empty()
        ]

    else:
        raise TypeError(
            'Unsupported extension: {}'.format(extension)
        )

    return BookAdapter(workbook, allsheet)
